#!/usr/bin/env python3
"""
Seed Database Generator for Elektropregled Android App

This script generates a prebuilt SQLite database file from:
1. Excel file (testniPodaci.xlsx) containing test data
2. PDF file (popisProvjeraPoVrstamaUredjaja.pdf) containing checklist items per device type

The generated database is placed in android-app/app/src/main/assets/database/seed.db
and is copied to the app's database directory on first run.

Usage:
    python tools/generate_seed_db.py

Requirements:
    pip install openpyxl PyPDF2 sqlite3
"""

import os
import sys
import sqlite3
import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import hashlib

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

try:
    import PyPDF2
except ImportError:
    print("WARNING: PyPDF2 not installed. PDF parsing will be skipped.")
    print("Install with: pip install PyPDF2")
    PyPDF2 = None


# Paths
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
INPUT_DIR = PROJECT_ROOT/ "baza" / "test-data"
EXCEL_FILE = INPUT_DIR / "testniPodaci.xlsx"
PDF_FILE = INPUT_DIR / "popisProvjeraPoVrstamaUredjaja.pdf"
OUTPUT_DIR = PROJECT_ROOT / "android-app" / "app" / "src" / "main" / "assets" / "database"
OUTPUT_DB = OUTPUT_DIR / "seed.db"
SCHEMA_SQL = PROJECT_ROOT / "baza" / "scripts" / "mobilna_sqlite.sql"
ROOM_SCHEMA_JSON = PROJECT_ROOT / "android-app" / "app" / "schemas" / "com.example.elektropregled.data.database.AppDatabase" / "1.json"

# Checklist mapping file (if PDF parsing is unreliable, use this JSON file)
CHECKLIST_MAPPING_FILE = INPUT_DIR / "checklist_mapping.json"


def create_database_schema(conn: sqlite3.Connection) -> None:
    """Create database schema using Room's exact schema from JSON."""
    print("Creating database schema from Room schema JSON...")
    
    if not ROOM_SCHEMA_JSON.exists():
        raise FileNotFoundError(f"Room schema JSON not found: {ROOM_SCHEMA_JSON}")
    
    # Load Room schema JSON
    with open(ROOM_SCHEMA_JSON, 'r', encoding='utf-8') as f:
        schema_data = json.load(f)
    
    # Enable foreign keys
    conn.execute("PRAGMA foreign_keys = ON")
    
    database = schema_data['database']
    
    # Create tables using Room's exact CREATE TABLE statements
    for entity in database['entities']:
        table_name = entity['tableName']
        create_sql = entity['createSql'].replace('${TABLE_NAME}', table_name)
        
        try:
            conn.execute(create_sql)
            print(f"  Created table: {table_name}")
        except sqlite3.Error as e:
            if "already exists" not in str(e).lower():
                print(f"Warning: Error creating table {table_name}: {e}")
                print(f"SQL: {create_sql[:200]}...")
    
    # Create indices using Room's exact CREATE INDEX statements
    for entity in database['entities']:
        table_name = entity['tableName']
        if 'indices' in entity:
            for index in entity['indices']:
                create_index_sql = index['createSql'].replace('${TABLE_NAME}', table_name)
                try:
                    conn.execute(create_index_sql)
                    print(f"  Created index: {index['name']}")
                except sqlite3.Error as e:
                    if "already exists" not in str(e).lower():
                        print(f"Warning: Error creating index {index['name']}: {e}")
    
    # Create room_master_table and set identity hash
    for setup_query in database['setupQueries']:
        try:
            conn.execute(setup_query)
        except sqlite3.Error as e:
            if "already exists" not in str(e).lower():
                print(f"Warning: Error executing setup query: {e}")
    
    # Set user_version to match Room database version
    conn.execute(f"PRAGMA user_version = {database['version']}")
    
    conn.commit()
    print("Schema created successfully using Room's exact schema.")


def load_excel_data(excel_file: Path) -> Dict[str, List[Dict]]:
    """Load data from Excel file."""
    print(f"Loading Excel data from {excel_file}...")
    
    if not excel_file.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_file}")
    
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    data = {}
    
    # Expected sheets
    expected_sheets = ['postrojenje', 'polje', 'vrstaUredjaja', 'uredjaj']
    
    for sheet_name in expected_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in Excel file")
            data[sheet_name] = []
            continue
        
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            print(f"Warning: Sheet '{sheet_name}' is empty")
            data[sheet_name] = []
            continue
        
        # First row is headers
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        print(f"  {sheet_name}: {len(rows) - 1} data rows, headers: {headers}")
        
        # Convert rows to dictionaries
        sheet_data = []
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue  # Skip empty rows
            
            row_dict = {}
            for i, header in enumerate(headers):
                value = row[i] if i < len(row) else None
                # Convert None to appropriate defaults
                if value is None:
                    row_dict[header] = None
                else:
                    row_dict[header] = value
            
            sheet_data.append(row_dict)
        
        data[sheet_name] = sheet_data
    
    wb.close()
    print(f"Loaded {sum(len(v) for v in data.values())} total records from Excel.")
    return data


def load_checklist_mapping() -> Dict[str, List[Dict]]:
    """Load checklist items mapping from JSON file or PDF."""
    print("Loading checklist items...")
    
    # Try JSON mapping file first (more reliable)
    if CHECKLIST_MAPPING_FILE.exists():
        print(f"  Loading from JSON mapping: {CHECKLIST_MAPPING_FILE}")
        with open(CHECKLIST_MAPPING_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    # Try PDF parsing
    if PDF_FILE.exists() and PyPDF2:
        print(f"  Attempting to parse PDF: {PDF_FILE}")
        try:
            return parse_pdf_checklist(PDF_FILE)
        except Exception as e:
            print(f"  Warning: PDF parsing failed: {e}")
            print(f"  Creating empty checklist mapping. Please create {CHECKLIST_MAPPING_FILE}")
    
    # Return empty mapping
    print("  No checklist mapping found. Using empty mapping.")
    return {}


def parse_pdf_checklist(pdf_file: Path) -> Dict[str, List[Dict]]:
    """Parse PDF to extract checklist items per device type."""
    # This is a basic implementation. For production, you may need more sophisticated PDF parsing.
    # For now, we'll create a template structure that can be filled manually.
    
    mapping = {}
    
    try:
        with open(pdf_file, 'rb') as f:
            pdf_reader = PyPDF2.PdfReader(f)
            
            # Extract text from all pages
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            
            # Basic parsing (this is a simplified version)
            # In production, you'd need to parse the PDF structure more carefully
            # For now, we'll return an empty structure and recommend using JSON mapping
            
            print("  Warning: PDF parsing is basic. Consider using checklist_mapping.json")
            
    except Exception as e:
        print(f"  Error parsing PDF: {e}")
    
    return mapping


def insert_postrojenja(conn: sqlite3.Connection, data: List[Dict]) -> None:
    """Insert postrojenja (facilities) into database."""
    print(f"Inserting {len(data)} postrojenja...")
    
    conn.execute("PRAGMA foreign_keys = ON")
    
    for row in data:
        try:
            conn.execute("""
                INSERT OR REPLACE INTO Postrojenje (id_postr, ozn_vr_postr, naz_postr, lokacija)
                VALUES (?, ?, ?, ?)
            """, (
                int(row.get('id_postr', 0)),
                str(row.get('ozn_vr_postr', '')),
                str(row.get('naz_postr', '')),
                str(row.get('lokacija', '')) if row.get('lokacija') else None
            ))
        except Exception as e:
            print(f"  Warning: Error inserting postrojenje {row.get('id_postr')}: {e}")
    
    conn.commit()
    print(f"  Inserted {len(data)} postrojenja.")


def insert_polja(conn: sqlite3.Connection, data: List[Dict]) -> None:
    """Insert polja (fields) into database."""
    print(f"Inserting {len(data)} polja...")
    
    conn.execute("PRAGMA foreign_keys = ON")
    
    for row in data:
        try:
            nap_razina = float(row.get('nap_razina', 0.0))
            conn.execute("""
                INSERT OR REPLACE INTO Polje (id_polje, nap_razina, ozn_vr_polje, naz_polje, id_postr)
                VALUES (?, ?, ?, ?, ?)
            """, (
                int(row.get('id_polje', 0)),
                nap_razina,
                str(row.get('ozn_vr_polje', '')),
                str(row.get('naz_polje', '')),
                int(row.get('id_postr', 0))
            ))
        except Exception as e:
            print(f"  Warning: Error inserting polje {row.get('id_polje')}: {e}")
    
    conn.commit()
    print(f"  Inserted {len(data)} polja.")


def insert_vrste_uredaja(conn: sqlite3.Connection, data: List[Dict], checklist_mapping: Dict) -> None:
    """Insert vrste uredaja (device types) and their checklist parameters."""
    print(f"Inserting {len(data)} vrste uredaja...")
    
    conn.execute("PRAGMA foreign_keys = ON")
    
    # Create a mapping from ozn_vr_ured to id_vr_ured
    ozn_to_id = {}
    next_id = 1
    
    for row in data:
        ozn = str(row.get('ozn_vr_ured', '')).strip()
        naz = str(row.get('naz_vr_ured', '')).strip()
        
        if not ozn:
            continue
        
        # Generate deterministic ID from ozn (hash-based)
        vrsta_id = abs(hash(ozn)) % (2**31 - 1)
        if vrsta_id == 0:
            vrsta_id = 1
        
        # Check for collisions
        while True:
            cursor = conn.execute("SELECT id_vr_ured FROM VrstaUredaja WHERE id_vr_ured = ?", (vrsta_id,))
            if cursor.fetchone() is None:
                break
            vrsta_id = (vrsta_id + 1) % (2**31 - 1)
        
        try:
            conn.execute("""
                INSERT OR REPLACE INTO VrstaUredaja (id_vr_ured, ozn_vr_ured, naz_vr_ured)
                VALUES (?, ?, ?)
            """, (vrsta_id, ozn, naz))
            
            ozn_to_id[ozn] = vrsta_id
        except Exception as e:
            print(f"  Warning: Error inserting vrsta uredaja {ozn}: {e}")
    
    conn.commit()
    print(f"  Inserted {len(ozn_to_id)} vrste uredaja.")
    
    # Insert checklist parameters for each device type
    print("Inserting checklist parameters...")
    parametar_id_counter = 1
    
    for ozn, vrsta_id in ozn_to_id.items():
        checklist_items = checklist_mapping.get(ozn, [])
        
        if not checklist_items:
            # Create default checklist items if none provided
            # This ensures every device type has at least basic parameters
            checklist_items = [
                {
                    "naz_parametra": "Vizualna provjera",
                    "tip_podataka": "BOOLEAN",
                    "redoslijed": 1,
                    "opis": "Vizualna provjera općeg stanja"
                },
                {
                    "naz_parametra": "Napomena",
                    "tip_podataka": "TEXT",
                    "redoslijed": 2,
                    "opis": "Dodatne napomene"
                }
            ]
        
        for item in checklist_items:
            try:
                # Generate deterministic parametar ID
                param_id = abs(hash(f"{ozn}_{item.get('naz_parametra', '')}")) % (2**31 - 1)
                if param_id == 0:
                    param_id = 1
                
                # Check for collisions
                while True:
                    cursor = conn.execute("SELECT id_parametra FROM ParametarProvjere WHERE id_parametra = ?", (param_id,))
                    if cursor.fetchone() is None:
                        break
                    param_id = (param_id + 1) % (2**31 - 1)
                
                tip_podataka = item.get('tip_podataka', 'TEXT').upper()
                min_val = item.get('min_vrijednost')
                max_val = item.get('max_vrijednost')
                mjerna_jedinica = item.get('mjerna_jedinica')
                
                # Validate tip_podataka
                if tip_podataka not in ['BOOLEAN', 'NUMERIC', 'TEXT']:
                    tip_podataka = 'TEXT'
                
                # For NUMERIC, ensure mjerna_jedinica is set
                if tip_podataka == 'NUMERIC' and not mjerna_jedinica:
                    mjerna_jedinica = '-'
                
                conn.execute("""
                    INSERT OR REPLACE INTO ParametarProvjere (
                        id_parametra, naz_parametra, tip_podataka, min_vrijednost,
                        max_vrijednost, mjerna_jedinica, obavezan, redoslijed, opis, id_vr_ured
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    param_id,
                    str(item.get('naz_parametra', '')),
                    tip_podataka,
                    float(min_val) if min_val is not None else None,
                    float(max_val) if max_val is not None else None,
                    mjerna_jedinica,
                    1 if item.get('obavezan', True) else 0,
                    int(item.get('redoslijed', parametar_id_counter)),
                    item.get('opis'),
                    vrsta_id
                ))
                
                parametar_id_counter += 1
            except Exception as e:
                print(f"  Warning: Error inserting parametar for {ozn}: {e}")
    
    conn.commit()
    print(f"  Inserted checklist parameters for {len(ozn_to_id)} device types.")


def insert_uredaji(conn: sqlite3.Connection, data: List[Dict], ozn_to_id_map: Dict[str, int]) -> None:
    """Insert uredaji (devices) into database."""
    print(f"Inserting {len(data)} uredaji...")
    
    conn.execute("PRAGMA foreign_keys = ON")
    
    inserted = 0
    for row in data:
        try:
            id_ured = int(row.get('id_ured', 0))
            ozn_vr_ured = str(row.get('ozn_vr_ured', '')).strip()
            id_vr_ured = ozn_to_id_map.get(ozn_vr_ured)
            
            if not id_vr_ured:
                print(f"  Warning: Unknown ozn_vr_ured '{ozn_vr_ured}' for uredaj {id_ured}, skipping")
                continue
            
            id_polje = row.get('id_polje')
            if id_polje is not None:
                try:
                    id_polje = int(id_polje) if id_polje != '' else None
                except (ValueError, TypeError):
                    id_polje = None
            
            conn.execute("""
                INSERT OR REPLACE INTO Uredaj (
                    id_ured, natp_plocica, tv_broj, id_postr, id_polje, id_vr_ured
                ) VALUES (?, ?, ?, ?, ?, ?)
            """, (
                id_ured,
                str(row.get('natp_plocica', '')),
                str(row.get('tv_broj', '')),
                int(row.get('id_postr', 0)),
                id_polje,
                id_vr_ured
            ))
            inserted += 1
        except Exception as e:
            print(f"  Warning: Error inserting uredaj {row.get('id_ured')}: {e}")
    
    conn.commit()
    print(f"  Inserted {inserted} uredaji.")


def validate_database(conn: sqlite3.Connection) -> None:
    """Validate database integrity."""
    print("Validating database...")
    
    conn.execute("PRAGMA foreign_keys = ON")
    
    # Check table counts
    tables = ['Postrojenje', 'Polje', 'VrstaUredaja', 'Uredaj', 'ParametarProvjere']
    for table in tables:
        cursor = conn.execute(f"SELECT COUNT(*) FROM {table}")
        count = cursor.fetchone()[0]
        print(f"  {table}: {count} records")
    
    # Check foreign key constraints
    cursor = conn.execute("PRAGMA foreign_key_check")
    issues = cursor.fetchall()
    if issues:
        print(f"  Warning: {len(issues)} foreign key constraint violations found")
        for issue in issues[:10]:  # Show first 10
            print(f"    {issue}")
    else:
        print("  Foreign key constraints: OK")
    
    # Check that each device type has checklist parameters
    cursor = conn.execute("""
        SELECT v.id_vr_ured, v.ozn_vr_ured, COUNT(p.id_parametra) as param_count
        FROM VrstaUredaja v
        LEFT JOIN ParametarProvjere p ON v.id_vr_ured = p.id_vr_ured
        GROUP BY v.id_vr_ured
        HAVING param_count = 0
    """)
    missing_params = cursor.fetchall()
    if missing_params:
        print(f"  Warning: {len(missing_params)} device types without checklist parameters")
        for vrsta in missing_params:
            print(f"    {vrsta[1]} (id: {vrsta[0]})")
    else:
        print("  All device types have checklist parameters")
    
    print("Validation complete.")


def get_ozn_to_id_map(conn: sqlite3.Connection) -> Dict[str, int]:
    """Get mapping from ozn_vr_ured to id_vr_ured."""
    cursor = conn.execute("SELECT ozn_vr_ured, id_vr_ured FROM VrstaUredaja")
    return {row[0]: row[1] for row in cursor.fetchall()}


def main():
    """Main function to generate seed database."""
    print("=" * 60)
    print("Elektropregled Seed Database Generator")
    print("=" * 60)
    
    # Check inputs
    if not EXCEL_FILE.exists():
        print(f"ERROR: Excel file not found: {EXCEL_FILE}")
        sys.exit(1)
    
    if not SCHEMA_SQL.exists():
        print(f"ERROR: Schema SQL file not found: {SCHEMA_SQL}")
        sys.exit(1)
    
    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Use a temporary file to avoid Windows file lock issues
    # Create database in a temp file first, then replace the original
    temp_db = OUTPUT_DB.with_suffix('.tmp')
    
    # Remove temp file if it exists
    if temp_db.exists():
        try:
            temp_db.unlink()
        except Exception:
            pass  # Ignore errors deleting temp file
    
    # Remove existing database if it exists (try, but don't fail if locked)
    use_temp_file = False
    if OUTPUT_DB.exists():
        print(f"Removing existing database: {OUTPUT_DB}")
        try:
            OUTPUT_DB.unlink()
        except PermissionError:
            print("WARNING: Cannot delete existing database (file is locked).")
            print("         This usually happens when Android Studio or emulator has the file open.")
            print("         Creating temporary database file instead...")
            print("         The script will try to replace it automatically when done.")
            use_temp_file = True
        except Exception as e:
            print(f"WARNING: Could not delete existing database: {e}")
            print("         Creating temporary database file instead...")
            use_temp_file = True
    
    # Create database in temp file if original is locked, otherwise use original path
    db_path = temp_db if use_temp_file else OUTPUT_DB
    print(f"\nCreating database: {db_path}")
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys = ON")
    
    try:
        # Create schema
        create_database_schema(conn)
        
        # Load Excel data
        excel_data = load_excel_data(EXCEL_FILE)
        
        # Load checklist mapping
        checklist_mapping = load_checklist_mapping()
        
        # Insert data in correct order (respecting foreign keys)
        print("\nInserting data...")
        insert_postrojenja(conn, excel_data.get('postrojenje', []))
        insert_polja(conn, excel_data.get('polje', []))
        insert_vrste_uredaja(conn, excel_data.get('vrstaUredjaja', []), checklist_mapping)
        
        # Get ozn to id mapping for uredaji
        ozn_to_id_map = get_ozn_to_id_map(conn)
        insert_uredaji(conn, excel_data.get('uredjaj', []), ozn_to_id_map)
        
        # Validate
        print("\n" + "=" * 60)
        validate_database(conn)
        
        # Set user version (for migrations) - must match Room database version
        # Room database version is 1, so set user_version to 1
        conn.execute("PRAGMA user_version = 1")
        conn.commit()
        
        # Close connection before trying to replace file
        conn.close()
        
        # If we created a temp file, try to replace the original
        if db_path == temp_db:
            print(f"\nAttempting to replace {OUTPUT_DB} with {temp_db}...")
            try:
                # Try to delete the locked file one more time
                if OUTPUT_DB.exists():
                    OUTPUT_DB.unlink()
                # Replace with temp file
                temp_db.replace(OUTPUT_DB)
                print(f"SUCCESS: Database replaced successfully!")
            except PermissionError:
                print(f"\nWARNING: Could not replace {OUTPUT_DB} (file is still locked).")
                print(f"         The new database has been created as: {temp_db}")
                print(f"         Please:")
                print(f"         1. Close Android Studio and any running emulators")
                print(f"         2. Manually delete {OUTPUT_DB}")
                print(f"         3. Rename {temp_db} to seed.db")
                print(f"         Or run this script again after closing Android Studio.")
            except Exception as e:
                print(f"\nWARNING: Could not replace database: {e}")
                print(f"         The new database has been created as: {temp_db}")
                print(f"         Please manually replace {OUTPUT_DB} with {temp_db}")
        
        # Determine final database path
        if db_path == temp_db:
            # Check if replacement was successful
            if OUTPUT_DB.exists():
                final_db = OUTPUT_DB
            else:
                final_db = temp_db
        else:
            final_db = OUTPUT_DB
        
        print("\n" + "=" * 60)
        print(f"SUCCESS: Seed database generated at {final_db}")
        print(f"Database size: {final_db.stat().st_size / 1024:.2f} KB")
        print("=" * 60)
        
    except Exception as e:
        print(f"\nERROR: Failed to generate database: {e}")
        import traceback
        traceback.print_exc()
        # Clean up temp file on error
        if temp_db.exists():
            try:
                temp_db.unlink()
            except:
                pass
        sys.exit(1)
    finally:
        if 'conn' in locals():
            try:
                conn.close()
            except:
                pass


if __name__ == "__main__":
    main()
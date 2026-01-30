#!/usr/bin/env python3
"""
Seed Database Generator for Elektropregled Android App

This script generates a prebuilt SQLite database file from:
1. Excel file (testniPodaci.xlsx) containing test data
2. PDF file (popisProvjeraPoVrstamaUredjaja.pdf) containing checklist items per device type

The generated database is placed in android-app/app/src/main/assets/database/seed.db
and is copied to the app's database directory on first run.

Usage:
    python tools/generate_seed_db.py

Requirements:
    pip install openpyxl PyPDF2 sqlite3
"""

import os
import sys
import sqlite3
import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import hashlib

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Install with: pip install openpyxl")
    sys.exit(1)

try:
    import PyPDF2
except ImportError:
    print("WARNING: PyPDF2 not installed. PDF parsing will be skipped.")
    print("Install with: pip install PyPDF2")
    PyPDF2 = None


# Paths
SCRIPT_DIR = Path(__file__).parent
PROJECT_ROOT = SCRIPT_DIR.parent
INPUT_DIR = SCRIPT_DIR / "input"
EXCEL_FILE = INPUT_DIR / "testniPodaci.xlsx"
PDF_FILE = INPUT_DIR / "popisProvjeraPoVrstamaUredjaja.pdf"
OUTPUT_DIR = PROJECT_ROOT / "android-app" / "app" / "src" / "main" / "assets" / "database"
OUTPUT_DB = OUTPUT_DIR / "seed.db"
SCHEMA_SQL = PROJECT_ROOT / "baza" / "scripts" / "mobilna_sqlite.sql"

# Checklist mapping file (if PDF parsing is unreliable, use this JSON file)
CHECKLIST_MAPPING_FILE = INPUT_DIR / "checklist_mapping.json"


def create_database_schema(conn: sqlite3.Connection) -> None:
    """Create database schema from SQL file."""
    print("Creating database schema...")
    
    with open(SCHEMA_SQL, 'r', encoding='utf-8') as f:
        schema_sql = f.read()
    
    # Enable foreign keys
    conn.execute("PRAGMA foreign_keys = ON")
    
    # Remove comments and clean up
    lines = []
    for line in schema_sql.split('\n'):
        # Remove inline comments (-- comment)
        if '--' in line:
            line = line[:line.index('--')]
        line = line.strip()
        if line:
            lines.append(line)
    
    # Join lines and split by semicolon (but be careful with CHECK constraints)
    full_sql = ' '.join(lines)
    
    # Split by semicolon, but preserve CHECK constraints
    # This is a simplified approach - for production, use a proper SQL parser
    statements = []
    current = []
    paren_depth = 0
    
    i = 0
    while i < len(full_sql):
        char = full_sql[i]
        current.append(char)
        
        if char == '(':
            paren_depth += 1
        elif char == ')':
            paren_depth -= 1
        elif char == ';' and paren_depth == 0:
            statement = ''.join(current).strip()
            if statement:
                statements.append(statement)
            current = []
        
        i += 1
    
    # If there's a remaining statement without semicolon
    if current:
        statement = ''.join(current).strip()
        if statement:
            statements.append(statement)
    
    # Execute all statements
    for statement in statements:
        try:
            conn.execute(statement)
        except sqlite3.Error as e:
            # Only print error if it's not about existing tables/indexes
            if "already exists" not in str(e).lower():
                print(f"Warning: Error executing statement: {e}")
                if len(statement) > 200:
                    print(f"Statement (first 200 chars): {statement[:200]}...")
                else:
                    print(f"Statement: {statement}")
    
    conn.commit()
    print("Schema created successfully.")


def load_excel_data(excel_file: Path) -> Dict[str, List[Dict]]:
    """Load data from Excel file."""
    print(f"Loading Excel data from {excel_file}...")
    
    if not excel_file.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_file}")
    
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    data = {}
    
    # Expected sheets
    expected_sheets = ['postrojenje', 'polje', 'vrstaUredjaja', 'uredjaj']
    
    for sheet_name in expected_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in Excel file")
            data[sheet_name] = []
            continue
        
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            print(f"Warning: Sheet '{sheet_name}' is empty")
            data[sheet_name] = []
            continue
        
        # First row is headers
        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        print(f"  {sheet_name}: {len(rows) - 1} data rows, headers: {headers}")
        
        # Convert rows to dictionaries
        sheet_data = []
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue  # Skip empty rows
            
            row_dict = {}
            for i, header in enumerate(headers):
                value = row[i] if i < len(row) else None
                # Convert None to appropriate defaults
                if value is None:
                    row_dict[header] = None
                else:
                    row_dict[header] = value
            
            sheet_data.append(row_dict)
        
        data[sheet_name] = sheet_data
    
    wb.close()
    print(f"Loaded {sum(len(v) for v in data.values())} total records from Excel.")
    return data


def load_checklist_mapping() -> Dict[str, List[Dict]]:
    """Load checklist items mapping from JSON file or PDF."""
    print("Loading checklist items...")
    
    # Try JSON mapping file first (more reliable)
    if CHECKLIST_MAPPING_FILE.exists():
        print(f"  Loading from JSON mapping: {CHECKLIST_MAPPING_FILE}")
        with open(CHECKLIST_MAPPING_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    
    # Try PDF parsing
    if PDF_FILE.exists() and PyPDF2:
        print(f"  Attempting to parse PDF: {PDF_FILE}")
        try:
            return parse_pdf_checklist(PDF_FILE)
        except Exception as e:
            print(f"  Warning: PDF parsing failed: {e}")
            print(f"  Creating empty checklist mapping. Please create {CHECKLIST_MAPPING_FILE}")
    
    # Return empty mapping
    print("  No checklist mapping found. Using empty mapping.")
    return {}


def parse_pdf_checklist(pdf_file: Path) -> Dict[str, List[Dict]]:
    """Parse PDF to extract checklist items per device type."""
    # This is a basic implementation. For production, you may need more sophisticated PDF parsing.
    # For now, we'll create a template structure that can be filled manually.
    
    mapping = {}
    
    try:
        with open(pdf_file, 'rb') as f:
            pdf_reader = PyPDF2.PdfReader(f)
            
            # Extract text from all pages
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
            
            # Basic parsing (this is a simplified version)
            # In production, you'd need to parse the PDF structure more carefully
            # For now, we'll return an empty structure and recommend using JSON mapping
            
            print("  Warning: PDF parsing is basic. Consider using checklist_mapping.json")
            
    except Exception as e:
        print(f"  Error parsing PDF: {e}")
    
    return mapping


def insert_postrojenja(conn: sqlite3.Connection, data: List[Dict]) -> None:
    """Insert postrojenja (facilities) into database."""
    print(f"Inserting {len(data)} postrojenja...")
    
    conn.execute("PRAGMA foreign_keys = ON")
    
    for row in data:
        try:
            conn.execute("""
                INSERT OR REPLACE INTO Postrojenje (id_postr, ozn_vr_postr, naz_postr, lokacija)
                VALUES (?, ?, ?, ?)
            """, (
                int(row.get('id_postr', 0)),
                str(row.get('ozn_vr_postr', '')),
                str(row.get('naz_postr', '')),
                str(row.get('lokacija', '')) if row.get('lokacija') else None
            ))
        except Exception as e:
            print(f"  Warning: Error inserting postrojenje {row.get('id_postr')}: {e}")
    
    conn.commit()
    print(f"  Inserted {len(data)} postrojenja.")


def insert_polja(conn: sqlite3.Connection, data: List[Dict]) -> None:
    """Insert polja (fields) into database."""
    print(f"Inserting {len(data)} polja...")
    
    conn.execute("PRAGMA foreign_keys = ON")
    
    for row in data:
        try:
            nap_razina = float(row.get('nap_razina', 0.0))
            conn.execute("""
                INSERT OR REPLACE INTO Polje (id_polje, nap_razina, ozn_vr_polje, naz_polje, id_postr)
                VALUES (?, ?, ?, ?, ?)
            """, (
                int(row.get('id_polje', 0)),
                nap_razina,
                str(row.get('ozn_vr_polje', '')),
                str(row.get('naz_polje', '')),
                int(row.get('id_postr', 0))
            ))
        except Exception as e:
            print(f"  Warning: Error inserting polje {row.get('id_polje')}: {e}")
    
    conn.commit()
    print(f"  Inserted {len(data)} polja.")


def insert_vrste_uredaja(conn: sqlite3.Connection, data: List[Dict], checklist_mapping: Dict) -> None:
    """Insert vrste uredaja (device types) and their checklist parameters."""
    print(f"Inserting {len(data)} vrste uredaja...")
    
    conn.execute("PRAGMA foreign_keys = ON")
    
    # Create a mapping from ozn_vr_ured to id_vr_ured
    ozn_to_id = {}
    next_id = 1
    
    for row in data:
        ozn = str(row.get('ozn_vr_ured', '')).strip()
        naz = str(row.get('naz_vr_ured', '')).strip()
        
        if not ozn:
            continue
        
        # Generate deterministic ID from ozn (hash-based)
        vrsta_id = abs(hash(ozn)) % (2**31 - 1)
        if vrsta_id == 0:
            vrsta_id = 1
        
        # Check for collisions
        while True:
            cursor = conn.execute("SELECT id_vr_ured FROM VrstaUredaja WHERE id_vr_ured = ?", (vrsta_id,))
            if cursor.fetchone() is None:
                break
            vrsta_id = (vrsta_id + 1) % (2**31 - 1)
        
        try:
            conn.execute("""
                INSERT OR REPLACE INTO VrstaUredaja (id_vr_ured, ozn_vr_ured, naz_vr_ured)
                VALUES (?, ?, ?)
            """, (vrsta_id, ozn, naz))
            
            ozn_to_id[ozn] = vrsta_id
        except Exception as e:
            print(f"  Warning: Error inserting vrsta uredaja {ozn}: {e}")
    
    conn.commit()
    print(f"  Inserted {len(ozn_to_id)} vrste uredaja.")
    
    # Insert checklist parameters for each device type
    print("Inserting checklist parameters...")
    parametar_id_counter = 1
    
    for ozn, vrsta_id in ozn_to_id.items():
        checklist_items = checklist_mapping.get(ozn, [])
        
        if not checklist_items:
            # Create default checklist items if none provided
            # This ensures every device type has at least basic parameters
            checklist_items = [
                {
                    "naz_parametra": "Vizualna provjera",
                    "tip_podataka": "BOOLEAN",
                    "obavezan": True,
                    "redoslijed": 1,
                    "opis": "Vizualna provjera općeg stanja"
                },
                {
                    "naz_parametra": "Napomena",
                    "tip_podataka": "TEXT",
                    "obavezan": False,
                    "redoslijed": 2,
                    "opis": "Dodatne napomene"
                }
            ]
        
        for item in checklist_items:
            try:
                # Generate deterministic parametar ID
                param_id = abs(hash(f"{ozn}_{item.get('naz_parametra', '')}")) % (2**31 - 1)
                if param_id == 0:
                    param_id = 1
                
                # Check for collisions
                while True:
                    cursor = conn.execute("SELECT id_parametra FROM ParametarProvjere WHERE id_parametra = ?", (param_id,))
                    if cursor.fetchone() is None:
                        break
                    param_id = (param_id + 1) % (2**31 - 1)
                
                tip_podataka = item.get('tip_podataka', 'TEXT').upper()
                min_val = item.get('min_vrijednost')
                max_val = item.get('max_vrijednost')
                mjerna_jedinica = item.get('mjerna_jedinica')
                
                # Validate tip_podataka
                if tip_podataka not in ['BOOLEAN', 'NUMERIC', 'TEXT']:
                    tip_podataka = 'TEXT'
                
                # For NUMERIC, ensure mjerna_jedinica is set
                if tip_podataka == 'NUMERIC' and not mjerna_jedinica:
                    mjerna_jedinica = '-'
                
                conn.execute("""
                    INSERT OR REPLACE INTO ParametarProvjere (
                        id_parametra, naz_parametra, tip_podataka, min_vrijednost,
                        max_vrijednost, mjerna_jedinica, obavezan, redoslijed, opis, id_vr_ured
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    param_id,
                    str(item.get('naz_parametra', '')),
                    tip_podataka,
                    float(min_val) if min_val is not None else None,
                    float(max_val) if max_val is not None else None,
                    mjerna_jedinica,
                    1 if item.get('obavezan', True) else 0,
                    int(item.get('redoslijed', parametar_id_counter)),
                    item.get('opis'),
                    vrsta_id
                ))
                
                parametar_id_counter += 1
            except Exception as e:
                print(f"  Warning: Error inserting parametar for {ozn}: {e}")
    
    conn.commit()
    print(f"  Inserted checklist parameters for {len(ozn_to_id)} device types.")


def insert_uredaji(conn: sqlite3.Connection, data: List[Dict], ozn_to_id_map: Dict[str, int]) -> None:
    """Insert uredaji (devices) into database."""
    print(f"Inserting {len(data)} uredaji...")
    
    conn.execute("PRAGMA foreign_keys = ON")
    
    inserted = 0
    for row in data:
        try:
            id_ured = int(row.get('id_ured', 0))
            ozn_vr_ured = str(row.get('ozn_vr_ured', '')).strip()
            id_vr_ured = ozn_to_id_map.get(ozn_vr_ured)
            
            if not id_vr_ured:
                print(f"  Warning: Unknown ozn_vr_ured '{ozn_vr_ured}' for uredaj {id_ured}, skipping")
                continue
            
            id_polje = row.get('id_polje')
            if id_polje is not None:
                try:
                    id_polje = int(id_polje) if id_polje != '' else None
                except (ValueError, TypeError):
                    id_polje = None
            
            conn.execute("""
                INSERT OR REPLACE INTO Uredaj (
                    id_ured, natp_plocica, tv_broj, id_postr, id_polje, id_vr_ured
                ) VALUES (?, ?, ?, ?, ?, ?)
            """, (
                id_ured,
                str(row.get('natp_plocica', '')),
                str(row.get('tv_broj', '')),
                int(row.get('id_postr', 0)),
                id_polje,
                id_vr_ured
            ))
            inserted += 1
        except Exception as e:
            print(f"  Warning: Error inserting uredaj {row.get('id_ured')}: {e}")
    
    conn.commit()
    print(f"  Inserted {inserted} uredaji.")


def validate_database(conn: sqlite3.Connection) -> None:
    """Validate database integrity."""
    print("Validating database...")
    
    conn.execute("PRAGMA foreign_keys = ON")
    
    # Check table counts
    tables = ['Postrojenje', 'Polje', 'VrstaUredaja', 'Uredaj', 'ParametarProvjere']
    for table in tables:
        cursor = conn.execute(f"SELECT COUNT(*) FROM {table}")
        count = cursor.fetchone()[0]
        print(f"  {table}: {count} records")
    
    # Check foreign key constraints
    cursor = conn.execute("PRAGMA foreign_key_check")
    issues = cursor.fetchall()
    if issues:
        print(f"  Warning: {len(issues)} foreign key constraint violations found")
        for issue in issues[:10]:  # Show first 10
            print(f"    {issue}")
    else:
        print("  Foreign key constraints: OK")
    
    # Check that each device type has checklist parameters
    cursor = conn.execute("""
        SELECT v.id_vr_ured, v.ozn_vr_ured, COUNT(p.id_parametra) as param_count
        FROM VrstaUredaja v
        LEFT JOIN ParametarProvjere p ON v.id_vr_ured = p.id_vr_ured
        GROUP BY v.id_vr_ured
        HAVING param_count = 0
    """)
    missing_params = cursor.fetchall()
    if missing_params:
        print(f"  Warning: {len(missing_params)} device types without checklist parameters")
        for vrsta in missing_params:
            print(f"    {vrsta[1]} (id: {vrsta[0]})")
    else:
        print("  All device types have checklist parameters")
    
    print("Validation complete.")


def get_ozn_to_id_map(conn: sqlite3.Connection) -> Dict[str, int]:
    """Get mapping from ozn_vr_ured to id_vr_ured."""
    cursor = conn.execute("SELECT ozn_vr_ured, id_vr_ured FROM VrstaUredaja")
    return {row[0]: row[1] for row in cursor.fetchall()}


def main():
    """Main function to generate seed database."""
    print("=" * 60)
    print("Elektropregled Seed Database Generator")
    print("=" * 60)
    
    # Check inputs
    if not EXCEL_FILE.exists():
        print(f"ERROR: Excel file not found: {EXCEL_FILE}")
        sys.exit(1)
    
    if not SCHEMA_SQL.exists():
        print(f"ERROR: Schema SQL file not found: {SCHEMA_SQL}")
        sys.exit(1)
    
    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Remove existing database if it exists
    if OUTPUT_DB.exists():
        print(f"Removing existing database: {OUTPUT_DB}")
        OUTPUT_DB.unlink()
    
    # Create database
    print(f"\nCreating database: {OUTPUT_DB}")
    conn = sqlite3.connect(str(OUTPUT_DB))
    conn.execute("PRAGMA foreign_keys = ON")
    
    try:
        # Create schema
        create_database_schema(conn)
        
        # Load Excel data
        excel_data = load_excel_data(EXCEL_FILE)
        
        # Load checklist mapping
        checklist_mapping = load_checklist_mapping()
        
        # Insert data in correct order (respecting foreign keys)
        print("\nInserting data...")
        insert_postrojenja(conn, excel_data.get('postrojenje', []))
        insert_polja(conn, excel_data.get('polje', []))
        insert_vrste_uredaja(conn, excel_data.get('vrstaUredjaja', []), checklist_mapping)
        
        # Get ozn to id mapping for uredaji
        ozn_to_id_map = get_ozn_to_id_map(conn)
        insert_uredaji(conn, excel_data.get('uredjaj', []), ozn_to_id_map)
        
        # Validate
        print("\n" + "=" * 60)
        validate_database(conn)
        
        # Set user version (for migrations) - must match Room database version
        conn.execute("PRAGMA user_version = 2")
        conn.commit()
        
        print("\n" + "=" * 60)
        print(f"SUCCESS: Seed database generated at {OUTPUT_DB}")
        print(f"Database size: {OUTPUT_DB.stat().st_size / 1024:.2f} KB")
        print("=" * 60)
        
    except Exception as e:
        print(f"\nERROR: Failed to generate database: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
